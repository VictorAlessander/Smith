# coding: UTF-8


try:
    import cria_planilha
    import time
    import re
    from bs4 import BeautifulSoup as BSoup
    import requests
    import telepot
    from openpyxl import Workbook, load_workbook

except Exception as e1:
    print(red + '[!] Ocorreu um erro: {}' .format(e1) + restore)


"""
O programa funciona somente no site da kabum, pelo menos por enquanto.
Sinta-se a vontade para alterá-lo conforme a sua necessidade.

"""

red = '\033[31m'
yellow = '\033[33m'
green = '\033[32m'
negrito = '\033[1m'
restore = '\033[0;0m'


print("""
 __           _ _   _     
/ _\_ __ ___ (_) |_| |__  
\ \| '_ ` _ \| | __| '_ \ 
_\ \ | | | | | | |_| | | |
\__/_| |_| |_|_|\__|_| |_|                    



+------------------+
|[S]art / Sair = 0 |
+------------------+

Digite help para mais informações.

""")


def enviar_dados(dados):
    # Aqui deve ser inserida a chave do seu bot e o id do grupo onde ele irá atuar
    config = {'bot_key' : 'bot_key', 'group_id' : int('group_id')}
    bot = telepot.Bot(config['bot_key'])
    group = config['group_id']

    informacao = "Produto: {}\nParcelado: {}\nA vista: {}" .format(
        dados[0], dados[1], dados[2])

    bot.sendMessage(group, informacao)

    # Exibe os resultados
    print(informacao)
    print(green + "[+] Mensagem enviada com sucesso!" + restore)
    print('-' * 100)


# Falta implementar a função, por enquanto ela apenas repassa os dados.
def avaliar(nome, preco, a_vista):
    dados = [nome]

    dados.append(preco)
    dados.append(a_vista)

    enviar_dados(dados)


def extrair():
    try:
        a1 = open(nome_arquivo, 'r')  # Lê o arquivo com os links

        # Laço para percorrer todos os links do arquivo
        for x in a1.readlines():

            links = x

            try:
                print(yellow + '[+] Verificando se a url está online...' + restore)
                test = requests.head(links)
                print(green + '{}\n{}' .format(test, links) + restore)

            except:
                print(red + '404 {}' .format(links) + restore)
                print('-' * 100)
                continue

            # Variáveis que receberão as strings brutas

            nome = ""
            preco_antigo = ""
            preco = ""
            a_vista = ""

            # Abre o link informado na variável link
            open_link = requests.get(links)

            # Faz o parser das informações no link
            soup = BSoup(open_link.content, "html.parser")

            results = []  # Array que vai receber as strings tratadas

            # Procura nas linhas em html as tags/classes e os
            # seus respectivos valores são atribuídos as variáveis

            w = soup.find('h1', attrs={'class': 'titulo_det'})

            if w:
                nome = w.text

            x = soup.find('div', attrs={'class': 'preco_antigo'})

            if x:
                preco_antigo = x.text

            y = soup.find('div', attrs={'class': 'preco_normal'})

            if y:
                preco = y.text

            z = soup.find('span', attrs={'class': 'preco_desconto'})

            if z:
                z = soup.find('span', attrs={'itemprop': 'offers'})
                a_vista = z.text

            # Retira os caracteres especificados (tratamento da string)
            filter1 = lambda *argv: [
                            re.sub('[\t\n Ddepor]', '', arg) for arg in argv]

            preco_antigo, preco, a_vista = filter1(
                preco_antigo, preco, a_vista)

            # Adiciona ao array results a string tratada

            results.append(nome)
            results.append(preco_antigo)
            results.append(preco)
            results.append(a_vista)

            #  print("Produto: {}\nPreço Anterior:{}\nPreço Atual: {}\nÀ Vista: {}" .format(nome, preco_antigo, preco, a_vista))

            ws.append(results)  # Escreve na tabela o conteúdo do array results

            wb.save(planilha)  # Salva o arquivo .xlsx

            avaliar(nome, preco, a_vista)

        a1.close()

    except Exception as e2:
        print(red + "[!] Erro: {}" .format(e2) + restore)
        pass


def main():

    global planilha, nome_aba, nome_arquivo, ws, wb

    planilha = str(
        input("Insira o nome da planilha (Não esqueça de colocar a extensão .xlsx)\n"))

    try:
        wb = load_workbook(planilha)  # Carrega um arquivo xlsx existente

    except FileNotFoundError:
        print(green + "[+] Criando planilha. . ." + restore)
        cria_planilha.Planilha(planilha).criar()
        wb = load_workbook(planilha)  # Carrega um arquivo xlsx existente

    opcao = str.lower(
        input("[*] Deseja criar uma nova aba para inserir os dados? [S/N/C]\n"))

    if opcao == 's':
        nome_aba = str(input("Digite um nome para a nova aba: "))
        ws = wb.create_sheet(nome_aba)

    elif opcao == 'n':
        nome_aba = str(input("Digite o nome da aba: "))
        ws = wb.get_sheet_by_name(nome_aba)

    else:
        print(yellow + "[!] Escrevendo em uma aba existente..." + restore)
        ws = wb.active

    nome_arquivo = str(
        input("Nome do arquivo com os links (Não esqueça de colocar a extensão .txt)\n"))
    print(green + '[+] Iniciando' + restore)

    print(yellow + negrito + '[+] Ctrl + C para parar o programa.' + restore)
    while True:
        extrair()
        time.sleep(3600)


if __name__ == "__main__":

    opcao = str.lower(input('>> '))

    if opcao == 'help' or opcao == 'h':
        print("""
O programa abrirá um arquivo de texto com os links referentes ao site e
fará a extração dos dados contidos nos links de acordo com os parâmetros
de extração (nome do produto, preço). Esses dados serão jogados em uma
planilha que o programa irá criar caso a planilha não exista.
    """)

    elif opcao == 's':
        main()

    elif opcao == '0':
        exit(0)

    else:
        exit(0)
