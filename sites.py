# coding: UTF-8


try:
	import cria_planilha
	import time
	from bs4 import BeautifulSoup as BSoup
	import requests
	from openpyxl import Workbook, load_workbook

except Exception as e1:
	print('[!] Ocorreu um erro: {}' .format(e1))


def extrair(nome):

	planilha = nome

	try:
		wb = load_workbook(planilha) # Carrega um arquivo xlsx existente

	except FileNotFoundError:
		print("[*] Criando planilha. . .")
		cria_planilha.Planilha(planilha).criar()
		wb = load_workbook(planilha) # Carrega um arquivo xlsx existente
	
	opcao = str.lower(input("[*] Deseja criar uma nova aba para inserir os dados? [S/N/C]\n"))

	if opcao == 's':
		nome_aba = str(input("Digite um nome para a nova aba: "))
		ws = wb.create_sheet(nome_aba)
	
	elif opcao == 'n':
		nome_aba = str(input("Digite o nome da aba: "))
		ws = wb.get_sheet_by_name(nome_aba)

	else:
		print("[!] Escrevendo em uma aba existente. . .")
		ws = wb.active

	nome_arquivo = str(input("Nome do arquivo com os links (Não esqueça de colocar a extensão .txt)\n"))

	a1 = open(nome_arquivo, 'r') # Lê o arquivo com os links

	try:
		for x in a1:

			links = x

			# Variáveis que receberão as strings brutas

			nome = ""
			preco_antigo = ""
			preco = ""
			a_vista = ""

			open_link = requests.get(links) # Abre o link informado na variável link

			soup = BSoup(open_link.content, "html.parser") # Faz o parser das informações no link

			results = [] # Array que vai receber as strings tratadas

			# Procura nas linhas em html as tags/classes e o seu valor atribui a string nome

			#for x in soup.find('h1', attrs={'class': 'titulo_det'}):
			#	results.append(x) # Adiciona os elementos a lista
				#nome = x.text

			# Procura nas linhas em html as tags/classes e o seu valor atribui a string preco_antigo

			#for x in soup.findAll('div', attrs={'class': 'preco_antigo'}):
				#results.append(x.text) # Adiciona os elementos a lista 
			#	preco_antigo = x.text


			#for x in soup.findAll('div', attrs={'class': 'preco_normal'}):
			#	preco = x.text

			# Procura nas linhas em html as tags/classes e o seu valor atribui a string a_vista

			#for x in soup.findAll('span', attrs={'class': 'preco_desconto'}):
				#results.append(x.text) # Adiciona os elementos a lista
			#	a_vista = x.text

			x = soup.find('h1', attrs={'class' : 'titulo_det'})

			if x:
				nome = x.text

			w = soup.find('div', attrs={'class' : 'preco_antigo'})

			if w:
				preco_antigo = w.text

			y = soup.find('div', attrs={'class' : 'preco_normal'})

			if y:
				preco = y.text

			z = soup.find('span', attrs={'class' : 'preco_desconto'})

			if z:
				a_vista = z.text


			#Retira os caracteres especificados (tratamento da string)

			preco_antigo = "".join(preco_antigo.split('\t'))
			preco_antigo = "".join(preco_antigo.split('\n'))
			preco = "".join(preco.split('\t'))
			preco = "".join(preco.split('\n'))
			a_vista = "".join(a_vista.split('\t'))
			a_vista = "".join(a_vista.split('\n'))

			#Adiciona ao array results a string tratada

			results.append(nome)
			results.append(preco_antigo)
			results.append(preco)
			results.append(a_vista)

			print(results) # Exibe no terminal o array results

			ws.append(results) # Escreve na tabela o conteúdo do array results

			wb.save(planilha) # Salva o arquivo .xlsx

			time.sleep(5)

	except Exception as e2:
		print("[!] Erro: {}" .format(e2))
		pass